from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.enums import AuditAction, IPAddressStatus, ImportJobStatus, SourceType
from app.models.ipam import IPAddress, IPAddressSourceObservation, SourceRecord, Subnet
from app.models.jobs import ImportJob
from app.services.audit import AuditService
from app.importers.base import BaseImporter


class ExcelImporter(BaseImporter):
    expected_columns = [
        "address",
        "subnet_cidr",
        "hostname",
        "mac_address",
        "status",
        "source_reference",
        "confidence_score",
        "information_outlet_code",
    ]

    source_name = "excel"

    def validate(self, payload: dict) -> dict:
        workbook_path = payload.get("workbook_path")
        if not workbook_path:
            return {"status": "invalid", "message": "workbook_path is required"}
        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        missing = [column for column in self.expected_columns if column not in headers]
        if missing:
            return {"status": "invalid", "missing_columns": missing}
        return {"status": "ok", "expected_columns": self.expected_columns}

    def run(self, payload: dict) -> dict:
        db: Session = payload["db"]
        job: ImportJob = payload["job"]
        workbook_path = payload["workbook_path"]

        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        index = {header: idx for idx, header in enumerate(headers)}
        audit = AuditService(db)

        processed = created = updated = failed = 0
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[index["address"]]:
                continue
            processed += 1
            address = str(row[index["address"]]).strip()
            subnet_cidr = str(row[index["subnet_cidr"]]).strip()
            status_text = str(row[index["status"]]).strip()
            subnet = db.query(Subnet).filter(Subnet.cidr == subnet_cidr).first()
            if not subnet:
                failed += 1
                continue
            try:
                status = IPAddressStatus(status_text)
            except ValueError:
                failed += 1
                continue

            source_record = SourceRecord(
                source_type=SourceType.EXCEL,
                source_name=workbook_path,
                source_reference=str(row[index["source_reference"]]).strip() if row[index["source_reference"]] else f"row-{row_number}",
                source_system="excel_demo",
                import_job_id=job.id,
                row_identifier=str(row_number),
                confidence_score=float(row[index["confidence_score"]]) if row[index["confidence_score"]] is not None else None,
                raw_payload=str(row),
            )
            db.add(source_record)
            db.flush()

            ip = db.query(IPAddress).filter(IPAddress.address == address).first()
            if ip:
                ip.subnet_id = subnet.id
                ip.hostname = str(row[index["hostname"]]).strip() if row[index["hostname"]] else None
                ip.mac_address = str(row[index["mac_address"]]).strip() if row[index["mac_address"]] else None
                ip.status = status
                ip.primary_source_record_id = source_record.id
                ip.confidence_score = source_record.confidence_score
                ip.source_summary = f"Imported from {workbook_path}"
                updated += 1
            else:
                ip = IPAddress(
                    address=address,
                    subnet_id=subnet.id,
                    hostname=str(row[index["hostname"]]).strip() if row[index["hostname"]] else None,
                    mac_address=str(row[index["mac_address"]]).strip() if row[index["mac_address"]] else None,
                    status=status,
                    primary_source_record_id=source_record.id,
                    confidence_score=source_record.confidence_score,
                    source_summary=f"Imported from {workbook_path}",
                    trace_reference=source_record.source_reference,
                )
                db.add(ip)
                db.flush()
                created += 1

            observation = IPAddressSourceObservation(
                ip_address_id=ip.id,
                source_record_id=source_record.id,
                observed_hostname=ip.hostname,
                observed_mac_address=ip.mac_address,
                observed_status=ip.status.value,
                confidence_score=source_record.confidence_score,
                observed_at=datetime.now(timezone.utc),
                is_primary=True,
                notes=str(row[index["information_outlet_code"]]).strip() if row[index["information_outlet_code"]] else None,
            )
            db.add(observation)
            audit.log_event(
                action=AuditAction.IMPORT,
                entity_type="IPAddress",
                entity_id=ip.id,
                source_type=SourceType.EXCEL,
                source_reference=source_record.source_reference,
                correlation_id=job.id,
                after={"address": ip.address, "status": ip.status.value},
                change_summary="Excel demo import wrote or updated IP evidence.",
            )

        job.status = ImportJobStatus.COMPLETED if failed == 0 else ImportJobStatus.PARTIAL
        job.summary = f"Processed {processed} rows from Excel demo file."
        job.processed_count = processed
        job.created_count = created
        job.updated_count = updated
        job.failed_count = failed
        db.commit()
        db.refresh(job)
        return {"status": job.status.value, "message": job.summary, "processed": processed}

from pathlib import Path

from openpyxl import Workbook


def generate_demo_excel(target_path: Path) -> Path:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ip_data"
    sheet.append(
        [
            "address",
            "subnet_cidr",
            "hostname",
            "mac_address",
            "status",
            "source_reference",
            "confidence_score",
            "information_outlet_code",
        ]
    )
    sheet.append(
        ["10.10.110.20", "10.10.110.0/24", "demo-available-01", "00:11:22:33:44:20", "available_candidate", "row-available", 0.95, "IO-10F-1001-A"]
    )
    sheet.append(
        ["10.10.110.21", "10.10.110.0/24", "demo-dhcp-01", "00:11:22:33:44:21", "dhcp_pool", "row-dhcp", 0.92, "IO-10F-1001-A"]
    )
    sheet.append(
        ["10.10.110.22", "10.10.110.0/24", "demo-allocated-01", "00:11:22:33:44:22", "allocated_confirmed", "row-allocated", 0.97, "IO-10F-1001-A"]
    )
    sheet.append(
        ["10.10.110.23", "10.10.110.0/24", "demo-conflict-01", "00:11:22:33:44:23", "conflict_suspected", "row-conflict", 0.75, "IO-10F-1001-A"]
    )
    workbook.save(target_path)
    return target_path


def generate_capacity_alert_excel(target_path: Path) -> Path:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "capacity_demo"
    sheet.append(
        [
            "address",
            "subnet_cidr",
            "hostname",
            "mac_address",
            "status",
            "source_reference",
            "confidence_score",
            "information_outlet_code",
        ]
    )
    for host_octet in range(50, 66):
        sheet.append(
            [
                f"10.10.120.{host_octet}",
                "10.10.120.0/24",
                f"ops-capacity-{host_octet}",
                f"00:22:33:44:55:{host_octet:02d}",
                "allocated_confirmed" if host_octet < 62 else "reserved",
                f"capacity-row-{host_octet}",
                0.88,
                "IO-11F-1102-A",
            ]
        )
    workbook.save(target_path)
    return target_path


def generate_branch_demo_excel(target_path: Path) -> Path:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "branch_demo"
    sheet.append(
        [
            "address",
            "subnet_cidr",
            "hostname",
            "mac_address",
            "status",
            "source_reference",
            "confidence_score",
            "information_outlet_code",
        ]
    )
    rows = [
        ["10.20.210.40", "10.20.210.0/24", "osaka-available-01", "00:11:22:44:66:40", "available_candidate", "osaka-row-available", 0.82, "IO-03F-0301-A"],
        ["10.20.210.41", "10.20.210.0/24", "osaka-conflict-01", "00:11:22:44:66:41", "conflict_suspected", "osaka-row-conflict", 0.63, "IO-03F-0301-A"],
        ["10.20.210.42", "10.20.210.0/24", "osaka-dhcp-01", "00:11:22:44:66:42", "dhcp_pool", "osaka-row-dhcp", 0.77, "IO-03F-0301-B"],
    ]
    for row in rows:
        sheet.append(row)
    workbook.save(target_path)
    return target_path


def generate_all_demo_excels(base_dir: Path) -> list[Path]:
    return [
        generate_demo_excel(base_dir / "demo_ips.xlsx"),
        generate_capacity_alert_excel(base_dir / "capacity_alert_ips.xlsx"),
        generate_branch_demo_excel(base_dir / "branch_demo_ips.xlsx"),
    ]


if __name__ == "__main__":
    generate_all_demo_excels(Path(__file__).resolve().parents[1] / "demo")
